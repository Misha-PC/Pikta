import openpyxl


def write_file(sheet_context: list, path=r"output", file_name="out"):
    workbook = openpyxl.Workbook()              # create workbook
    workbook.remove(workbook.active)            # remove default worksheet
    for context in sheet_context:
        sheet = workbook.create_sheet()         # create new sheet
        for item, i in zip(context.items(), range(len(context.keys()))):
            key, val = item
            sheet.cell(1, i+1).value = key      # write headers
            indexs = sorted(val.keys())         # sort values by the "Y" component

            for index, serial in zip(indexs, range(len(indexs))):
                sheet.cell(serial + 2, i+1).value = val[index]  # write values

    file = f".\\{path}\\{file_name}.xlsx"
    print(f"Save: {file}")
    workbook.save(file)    # save workbook to file


